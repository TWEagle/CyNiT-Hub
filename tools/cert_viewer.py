
# tools/cert_viewer.py
# !/usr/bin/env python3
"""
Certificate / CSR Viewer voor CyNiT Hub (content-only, hub layout levert header/footer/CSS/JS).
- Route: GET/POST /cert  (decoderen en tonen)
- Exports:
  GET /cert/download/<fmt>?token=...   (fmt: json|csv|html|md|xlsx)
  GET /cert/download/zip_all?token=...
  GET /cert/save_md?token=...          (schrijft naar ./exports)

Afhankelijkheden:
- cryptography  (decoderen van CERT/CSR; PEM/DER)
- openpyxl (optioneel; enkel vereist voor .xlsx export)
"""

from __future__ import annotations

import base64
import io
import json
import re
import zipfile
import secrets
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Optional, Tuple, List

from flask import Flask, request, render_template_string, make_response, send_file, abort

# Gebruik jouw centrale hub layout
from beheer.main_layout import render_page as hub_render_page  # type: ignore

# ===== Paths / opslag =====
BASE_DIR = Path(__file__).resolve().parents[1]  # CyNiT-Hub/
EXPORTS_DIR = BASE_DIR / "exports"
TMP_DIR = BASE_DIR / "tmp" / "cert_viewer"
TMP_DIR.mkdir(parents=True, exist_ok=True)

# In-memory store: token -> info dict
_STORE: Dict[str, Dict[str, Any]] = {}

# ====== Helpers voor detectie & parsing ======
_B64_RE = re.compile(r"^[A-Za-z0-9+/=\s]+$")


def _now_utc() -> datetime:
    return datetime.now(timezone.utc)


def _strip_xml_wrapper(s: str) -> str:
    # Soms zijn certs geplakt uit XML; verwijder tags
    return re.sub(r"<[^>]+>", "", s)


def _try_base64_to_der_bytes(text: str) -> Optional[bytes]:
    if not text:
        return None
    t = _strip_xml_wrapper(text.strip()).strip()
    if "BEGIN " in t or "END " in t:
        return None
    if not _B64_RE.match(t):
        return None
    b64 = "".join(t.split())
    try:
        return base64.b64decode(b64, validate=False)
    except Exception:
        return None


def _normalize_pem(text: str) -> str:
    lines = [ln.strip() for ln in (text or "").replace("\r\n", "\n").split("\n") if ln.strip()]
    return "\n".join(lines) + ("\n" if lines else "")


def load_cert_or_csr(data: bytes) -> Tuple[str, Any]:
    """
    Detecteer CERT/CSR en PEM/DER.
    Return ("cert", x509.Certificate) of ("csr", x509.CertificateSigningRequest).
    """
    from cryptography import x509

    text: Optional[str]
    try:
        text = data.decode("utf-8", errors="ignore")
    except Exception:
        text = None

    if text:
        stripped = text.strip()
        # PEM CSR
        if "BEGIN CERTIFICATE REQUEST" in stripped or "BEGIN NEW CERTIFICATE REQUEST" in stripped:
            norm = _normalize_pem(stripped)
            try:
                csr = x509.load_pem_x509_csr(norm.encode("ascii", errors="ignore"))
                return "csr", csr
            except Exception:
                csr = x509.load_pem_x509_csr(data)
                return "csr", csr
        # PEM CERT
        if "BEGIN CERTIFICATE" in stripped and "REQUEST" not in stripped:
            cert = x509.load_pem_x509_certificate(data)
            return "cert", cert
        # plain base64 (DER)
        der = _try_base64_to_der_bytes(stripped)
        if der:
            try:
                cert = x509.load_der_x509_certificate(der)
                return "cert", cert
            except Exception:
                csr = x509.load_der_x509_csr(der)
                return "csr", csr

    # DER CERT
    try:
        cert = x509.load_der_x509_certificate(data)
        return "cert", cert
    except Exception:
        pass

    # DER CSR
    csr = x509.load_der_x509_csr(data)
    return "csr", csr


def _name_to_dict(name) -> Dict[str, str]:
    out: Dict[str, str] = {}
    for attr in name:
        key = getattr(attr.oid, "_name", None) or attr.oid.dotted_string
        out[key] = attr.value
    return out


def _pubkey_summary(pub) -> str:
    try:
        from cryptography.hazmat.primitives.asymmetric import rsa, dsa, ec

        if isinstance(pub, rsa.RSAPublicKey):
            return f"RSA {pub.key_size} bits"
        if isinstance(pub, dsa.DSAPublicKey):
            return f"DSA {pub.key_size} bits"
        if isinstance(pub, ec.EllipticCurvePublicKey):
            return f"EC {pub.curve.name}"
        return pub.__class__.__name__
    except Exception:
        return "unknown"


def _hash_name(sig_hash) -> str:
    try:
        return sig_hash.name
    except Exception:
        return "unknown"


def decode_cert_from_bytes(data: bytes, filename: str = "input") -> Dict[str, Any]:
    from cryptography import x509

    kind, obj = load_cert_or_csr(data)
    info: Dict[str, Any] = {
        "kind": kind,
        "filename": filename,
        "decoded_at_utc": _now_utc().isoformat(),
        "subject": {},
        "issuer": {},
        "properties": {},
        "extensions": [],
        "checks": [],
    }

    if kind == "cert":
        cert: x509.Certificate = obj
        info["subject"] = _name_to_dict(cert.subject)
        info["issuer"] = _name_to_dict(cert.issuer)

        props = info["properties"]
        props["serial_number"] = str(cert.serial_number)
        # Normaliseer naar UTC
        props["not_valid_before_utc"] = cert.not_valid_before.replace(tzinfo=timezone.utc).isoformat()
        props["not_valid_after_utc"] = cert.not_valid_after.replace(tzinfo=timezone.utc).isoformat()
        props["signature_hash"] = _hash_name(cert.signature_hash_algorithm)
        props["public_key"] = _pubkey_summary(cert.public_key())

        # Basischecks
        now = _now_utc()
        if cert.not_valid_before.replace(tzinfo=timezone.utc) > now:
            info["checks"].append({"name": "validity", "status": "WARN", "message": "Cert is nog niet geldig (not_before in de toekomst)."})
        if cert.not_valid_after.replace(tzinfo=timezone.utc) < now:
            info["checks"].append({"name": "validity", "status": "FAIL", "message": "Cert is verlopen (not_after in het verleden)."})

        if not info["checks"]:
            info["checks"].append({"name": "validity", "status": "OK", "message": "Validity window OK t.o.v. huidige UTC tijd."})

        for ext in cert.extensions:
            info["extensions"].append({
                "oid": ext.oid.dotted_string,
                "name": getattr(ext.oid, "_name", None) or "extension",
                "critical": bool(ext.critical),
                "value": str(ext.value)[:8000],
            })

    else:
        csr: x509.CertificateSigningRequest = obj
        info["subject"] = _name_to_dict(csr.subject)
        info["issuer"] = {}  # CSR heeft geen issuer
        props = info["properties"]
        props["signature_hash"] = _hash_name(csr.signature_hash_algorithm)
        props["public_key"] = _pubkey_summary(csr.public_key())
        for ext in csr.extensions:
            info["extensions"].append({
                "oid": ext.oid.dotted_string,
                "name": getattr(ext.oid, "_name", None) or "extension",
                "critical": bool(ext.critical),
                "value": str(ext.value)[:8000],
            })
        info["checks"].append({"name": "issuer", "status": "INFO", "message": "CSR heeft geen issuer (wordt pas ingevuld bij issuance)."})

    return info


# ======= CONTENT-ONLY TEMPLATE (layout verzorgt header/footer) =======
CONTENT_TEMPLATE = r"""
<style>
/* Compacte kaartjes; rest komt uit /static/main.css */
.certwrap { max-width: 1100px; margin: 0 auto; }
.card {
  background: rgba(10,15,18,.85); border-radius: 12px; padding: 16px 20px; margin-bottom: 20px;
  border: 1px solid var(--border, rgba(255,255,255,.10));
}
.card h2 { margin: 0 0 8px 0; }
.card small { color: var(--muted, #9fb3b3); }
.field-row { margin-bottom: 10px; }
.field-row label { display: block; margin-bottom: 3px; }
.in {
  width: 100%; box-sizing: border-box; padding: 10px 12px;
  border-radius: 10px; background: #111; color: #fff; border: 1px solid var(--border, rgba(255,255,255,.18));
}
textarea.in { min-height: 120px; resize: vertical; }
.row-inline { display: flex; gap: 12px; }
.row-inline > div { flex: 1; }
.error-box {
  background: #330000; border: 1px solid #aa3333; color: #ffaaaa; padding: 8px 10px;
  border-radius: 8px; margin-bottom: 12px; font-size: 0.9rem; white-space: pre-wrap;
}
.tbl { width: 100%; border-collapse: collapse; }
.tbl th, .tbl td { border-bottom: 1px solid rgba(255,255,255,.10); padding: 8px 8px; text-align: left; }
.tbl th { color: var(--muted, #9fb3b3); font-weight: 700; }
.kv { white-space: pre-wrap; word-break: break-word; }
</style>

<div class="certwrap">
  <h1>Certificate / CSR Viewer</h1>
  <p class="muted">Upload een certificaat/CSR (PEM/DER) of plak PEM/Base64. Daarna kan je exporteren naar JSON/CSV/HTML/MD/XLSX of alles in één ZIP.</p>

  {% if error %}
    <div class="error-box">{{ error }}</div>
  {% endif %}

  <div class="card">
    <h2>Input</h2>
    <form method="post" action="/cert" enctype="multipart/form-data">
      <div class="row-inline">
        <div>
          <label>Bestand upload</label>
          <input class="in" type="file" name="file" accept=".cer,.crt,.pem,.der,.csr,.txt">
        </div>
      </div>
      <div class="field-row">
        <label>Of plak PEM / Base64</label>
        <textarea class="in" name="pasted" placeholder="-----BEGIN CERTIFICATE----- ..."></textarea>
      </div>
      <div class="field-row">
        <button class="btn" type="submit">Decode</button>
      </div>
    </form>
  </div>

  {% if info and token %}
  <div class="card">
    <h2>Exports</h2>
    <p>
      <a class="btn" href="/cert/download/json?token={{ token }}">JSON</a>
      <a class="btn" href="/cert/download/csv?token={{ token }}">CSV</a>
      <a class="btn" href="/cert/download/xlsx?token={{ token }}">XLSX</a>
      <a class="btn" href="/cert/download/html?token={{ token }}">HTML</a>
      <a class="btn" href="/cert/download/md?token={{ token }}">MD</a>
      <a class="btn" href="/cert/download/zip_all?token={{ token }}">ZIP (alles)</a>
      <a class="btn" href="/cert/save_md?token={{ token }}">Save MD → exports/</a>
    </p>
  </div>

  <div class="card">
    <h2>Decoded</h2>
    <p class="muted">{{ info.kind }} · {{ info.filename }}</p>

    {% if info.checks %}
      <h3>Checks</h3>
      <table class="tbl">
        {% for c in info.checks %}
          <tr><th class="kv">{{ c.name }}</th><td class="kv">{{ c.status }} · {{ c.message }}</td></tr>
        {% endfor %}
      </table>
    {% endif %}

    <h3>Subject</h3>
    <table class="tbl">
      {% for k, v in (info.subject or {}).items() %}
        <tr><th class="kv">{{ k }}</th><td class="kv">{{ v }}</td></tr>
      {% endfor %}
    </table>

    <h3>Issuer</h3>
    {% if info.issuer %}
      <table class="tbl">
        {% for k, v in (info.issuer or {}).items() %}
          <tr><th class="kv">{{ k }}</th><td class="kv">{{ v }}</td></tr>
        {% endfor %}
      </table>
    {% else %}
      <p class="muted">CSR heeft geen issuer; dit wordt pas ingevuld na uitgifte van het certificaat.</p>
    {% endif %}

    <h3>Properties</h3>
    <table class="tbl">
      {% for k, v in (info.properties or {}).items() %}
        <tr><th class="kv">{{ k }}</th><td class="kv">{{ v }}</td></tr>
      {% endfor %}
    </table>

    {% if info.extensions %}
      <h3>Extensions</h3>
      <table class="tbl">
        {% for e in info.extensions %}
          <tr><th class="kv">{{ e.name }}</th><td class="kv">OID: {{ e.oid }}{% if e.critical %} · <strong>critical</strong>{% endif %}<br>{{ e.value }}</td></tr>
        {% endfor %}
      </table>
    {% endif %}
  </div>
  {% endif %}
</div>
"""


def _render_page(*, error: Optional[str] = None, info: Optional[Dict[str, Any]] = None, token: Optional[str] = None):
    content_html = render_template_string(CONTENT_TEMPLATE, error=error, info=info, token=token)
    return hub_render_page(title="Certificate / CSR Viewer", content_html=content_html)


# ======= Exports (helpers) =======
def build_csv_text(info: Dict[str, Any]) -> str:
    # Eenvoudige CSV (Section;Field;Value)
    lines = ["Section;Field;Value"]
    def add_section(sec: str, d: Dict[str, Any]):
        for k, v in (d or {}).items():
            lines.append(f"{sec};{k};{str(v).replace(';', ',')}")
    add_section("subject", info.get("subject", {}))
    add_section("issuer", info.get("issuer", {}))
    add_section("properties", info.get("properties", {}))
    for e in info.get("checks", []):
        lines.append(f"checks;{e.get('name','')};{e.get('status','')} - {e.get('message','')}")
    for e in info.get("extensions", []):
        nm = e.get("name", "")
        lines.append(f"extensions;{nm};OID={e.get('oid','')}, critical={e.get('critical', False)}")
    return "\n".join(lines)


def build_markdown(info: Dict[str, Any]) -> str:
    return (
        f"# {info.get('kind','cert').upper()} · {info.get('filename','')}\n\n"
        "## Checks\n" +
        "\n".join([f"- **{c.get('status','')}** — {c.get('name','')}: {c.get('message','')}" for c in info.get("checks", [])]) + "\n\n" +
        "## Subject\n" +
        "\n".join([f"- {k}: `{v}`" for k, v in (info.get("subject") or {}).items()]) + "\n\n" +
        "## Issuer\n" +
        ("\n".join([f"- {k}: `{v}`" for k, v in (info.get("issuer") or {}).items()]) or "_(CSR heeft geen issuer)_") + "\n\n" +
        "## Properties\n" +
        "\n".join([f"- {k}: `{v}`" for k, v in (info.get("properties") or {}).items()]) + "\n\n" +
        "## Extensions\n" +
        "\n".join([f"- {e.get('name','extension')} · OID `{e.get('oid','')}`" + (" · **critical**" if e.get("critical") else "") for e in info.get("extensions", [])]) + "\n"
    )


def build_html(info: Dict[str, Any]) -> str:
    # Minimal, bruikbaar in e-mail/preview
    rows = "".join([f"<tr><th>{k}</th><td>{v}</td></tr>" for k, v in (info.get("properties") or {}).items()])
    return f"""<!doctype html>
<html><head><meta charset="utf-8"><title>Certificate/CSR</title>
<style>body{{font-family:Segoe UI,Arial,sans-serif;padding:18px;background:#0b0b0b;color:#ddd}}
table{{border-collapse:collapse}}th,td{{border-bottom:1px solid #333;padding:6px 10px;text-align:left}}
h2{{color:#9fb3b3}}</style></head>
<body>
<h1>{info.get('kind','cert').upper()} · {info.get('filename','')}</h1>
<h2>Properties</h2><table>{rows}</table>
</body></html>"""



# ===== Excel sheetnaam helper =====
_WS_INVALID = re.compile(r'[:\\/?*\[\]]')  # Verboden tekens in Excel-werkbladnamen

def _safe_ws_title(title: str, fallback: str = "Sheet1") -> str:
    """
    Maak een veilige werkbladnaam:
    - vervang verboden tekens door spatie
    - normaliseer whitespace
    - truncate tot 31 tekens
    - val terug op 'fallback' als het leeg wordt
    """
    t = _WS_INVALID.sub(" ", (title or "").strip())
    t = re.sub(r"\s+", " ", t)
    if not t:
        t = fallback
    return t[:31]


def build_xlsx_bytes(info: Dict[str, Any]) -> Optional[bytes]:
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except Exception:
        return None

    wb = Workbook()
    ws = wb.active

    # ✨ veilige, korte titel
    base = (info.get("kind", "cert").upper() + " " +
            (Path(info.get("filename", "document")).stem or "Document"))
    ws.title = _safe_ws_title(base, fallback="Certificate CSR")

    ws.append(["Section", "Field", "Value"])

    def add_rows(sec: str, d: Dict[str, Any]):
        for k, v in (d or {}).items():
            ws.append([sec, k, str(v)])

    add_rows("subject", info.get("subject", {}))
    add_rows("issuer", info.get("issuer", {}))
    add_rows("properties", info.get("properties", {}))

    for e in info.get("checks", []):
        ws.append(["checks", e.get("name", ""), f"{e.get('status','')} - {e.get('message','')}"])

    for e in info.get("extensions", []):
        ws.append(["extensions", e.get("name", ""), f"OID={e.get('oid','')}, critical={e.get('critical', False)}"])

    # simpele auto-width
    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(12, min(80, max_len + 2))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def _require_token() -> Dict[str, Any]:
    token = (request.args.get("token") or "").strip()
    if not token or token not in _STORE:
        abort(400, "Token ontbreekt of is ongeldig.")
    return _STORE[token]


# ======= Routes =======
def register_web_routes(app: Flask):
    @app.route("/cert", methods=["GET", "POST"])
    def cert_index():
        if request.method == "GET":
            return _render_page()

        error: Optional[str] = None
        info: Optional[Dict[str, Any]] = None

        up = request.files.get("file")
        pasted = (request.form.get("pasted") or "").strip()

        try:
            if up and up.filename:
                data = up.read()
                info = decode_cert_from_bytes(data, filename=up.filename)
            elif pasted:
                if "BEGIN " in pasted:
                    data = _normalize_pem(pasted).encode("utf-8", errors="ignore")
                else:
                    der = _try_base64_to_der_bytes(pasted)
                    if not der:
                        raise ValueError("Kon pasted input niet herkennen als PEM of Base64 DER.")
                    data = der
                info = decode_cert_from_bytes(data, filename="pasted.txt")
            else:
                raise ValueError("Geen bestand gekozen en niets geplakt.")
        except Exception as e:
            error = f"Fout bij decoderen: {e}"

        token = None
        if info:
            token = secrets.token_urlsafe(16)
            _STORE[token] = info

        return _render_page(error=error, info=info, token=token)

    @app.get("/cert/download/<fmt>")
    def cert_download(fmt: str):
        info = _require_token()
        base_name = Path(info.get("filename", "certificate")).stem or "certificate"
        fmt = (fmt or "").lower().strip()

        if fmt == "json":
            content = json.dumps(info, indent=2, ensure_ascii=False)
            return make_response((content, 200, {"Content-Type": "application/json; charset=utf-8",
                                                 "Content-Disposition": f'attachment; filename="{base_name}.json"'}))

        if fmt == "csv":
            content = build_csv_text(info)
            return make_response((content, 200, {"Content-Type": "text/csv; charset=utf-8",
                                                 "Content-Disposition": f'attachment; filename="{base_name}.csv"'}))

        if fmt == "html":
            content = build_html(info)
            return make_response((content, 200, {"Content-Type": "text/html; charset=utf-8",
                                                 "Content-Disposition": f'attachment; filename="{base_name}.html"'}))

        if fmt == "md":
            content = build_markdown(info)
            return make_response((content, 200, {"Content-Type": "text/markdown; charset=utf-8",
                                                 "Content-Disposition": f'attachment; filename="{base_name}.md"'}))

        if fmt == "xlsx":
            data = build_xlsx_bytes(info)
            if data is None:
                return make_response(("XLSX export vereist 'openpyxl'. Installeer: pip install openpyxl", 500))
            buf = io.BytesIO(data)
            buf.seek(0)
            return send_file(buf, as_attachment=True, download_name=f"{base_name}.xlsx",
                             mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        return make_response(("Onbekend exporttype.", 400))

    @app.get("/cert/download/zip_all")
    def cert_zip_all():
        info = _require_token()
        base_name = Path(info.get("filename", "certificate")).stem or "certificate"

        json_txt = json.dumps(info, indent=2, ensure_ascii=False)
        csv_txt = build_csv_text(info)
        md_txt = build_markdown(info)
        html_txt = build_html(info)
        xlsx_bytes = build_xlsx_bytes(info)  # kan None zijn

        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
            zf.writestr(f"{base_name}.json", json_txt.encode("utf-8"))
            zf.writestr(f"{base_name}.csv", csv_txt.encode("utf-8"))
            zf.writestr(f"{base_name}.md", md_txt.encode("utf-8"))
            zf.writestr(f"{base_name}.html", html_txt.encode("utf-8"))
            if xlsx_bytes:
                zf.writestr(f"{base_name}.xlsx", xlsx_bytes)

        buf.seek(0)
        return send_file(buf, as_attachment=True, download_name=f"{base_name}_all.zip", mimetype="application/zip")

    @app.get("/cert/save_md")
    def cert_save_md():
        info = _require_token()
        EXPORTS_DIR.mkdir(parents=True, exist_ok=True)

        slug = re.sub(r"[^A-Za-z0-9._-]+", "_", Path(info.get("filename", "certificate")).stem or "certificate")
        ts = datetime.now().strftime("%Y%m%d-%H%M%S")
        filename = f"{slug}_{ts}.md"
        dest = EXPORTS_DIR / filename

        md_txt = build_markdown(info)
        dest.write_text(md_txt, encoding="utf-8")

        msg_html = f"""
        <!doctype html><meta charset="utf-8">
        <title>Saved</title>
        <style>body{{font-family:Segoe UI,Arial,sans-serif;background:#0b0b0b;color:#ddd;padding:18px}}</style>
        <h2>Markdown opgeslagen</h2>
        <p><strong>{filename}</strong> opgeslagen in <code>exports/</code>.</p>
        <p><a href="/cert">Terug</a></p>
        """
        return make_response(msg_html, 200)


# Standalone test (optioneel)
if __name__ == "__main__":
    app = Flask(__name__)
    register_web_routes(app)
    app.run("127.0.0.1", 5001, debug=True)
